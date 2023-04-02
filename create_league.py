import openpyxl
import os
import shutil
from openpyxl.styles import Font, PatternFill

#output_dir = r"C:\Program Files (x86)\Steam\steamapps\common\rFactor 2\UserData\player\Settings"
output_dir = r"G:/dev/tmp/rf2"

template = """//[[gMa1.002f (c)2016 ]] [[ ]]
GT3
{{
{full_name}
{{
Team = {team}
Component = {aaComponent}
Skin = {aaSkin}
VehFile = {aaVehFile}
Description = {full_name}
Number = {number}
Classes = GT3 my_gt3_league
Category = my_gt3_league
Aggression = {aggression}
Reputation = {reputation}
Courtesy = {courtesy}
Composure = {composure}
Speed = {speed}
QualifySpeed = {qualify_speed}
WetSpeed = {wet_speed}
StartSkill = {start_skill}
Crash = {crash}
Recovery = {recovery}
CompletedLaps = {completed_laps}
MinRacingSkill = {min_racing_skill}
}}
}}"""

def getLiveryFiles(vehicle_name):
    workbook = openpyxl.load_workbook('data.xlsx', data_only=True)
    ws = workbook['Liveries']
    # iterate over rows starting from the second row (skipping the header)
    for row in ws.iter_rows(min_row=2):
        # check if the first column equals to the vehicle string
        if row[0].value == vehicle_name:
            # return the value from the second column
            return({
                "livery_name":row[0].value,
                "folder":row[1].value,
                "veh_file":row[2].value,
                })
            break  # stop iterating after the first match is found

def findJsonFile(json_file):
         for subdir, dirs, files in os.walk('./data'):
            for file in files:
                if file == json_file:
                    # If the file is found, return the path of the directory that contains it
                    folder_path = os.path.join(os.getcwd(), subdir)
                    return(folder_path)
                
def getBaseName(vehFile):
    with open(vehFile, 'r') as file:
        for line in file:
            if line.startswith('DefaultLivery='):
                value = line.split("=")[1].strip()
                stripped_value = value.strip('"')           
                return stripped_value.replace(".dds","")
            
    
def createRoster():
    workbook = openpyxl.load_workbook('data.xlsx', data_only=True)
    # Get the worksheet by name
    worksheet = workbook['League']

    # Create an empty dictionary to store the data
    data = []

    # Get the header row and convert it to a list
    header = [cell.value for cell in next(worksheet.iter_rows(min_row=1))]

    # Loop through each row in the worksheet, starting from the second row
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        # Create a dictionary for the row using the header as keys
        row_dict = dict(zip(header, row))
        # Add the row dictionary to the data dictionary using the team name as key
        data.append(row_dict)


    counter = {}
    for row in data:
        if row['vehicle'] is not None:
            vehicle_file = getLiveryFiles(row['vehicle'])
            vehicle_base_name = getBaseName("./data/"+vehicle_file['folder']+"/"+vehicle_file['veh_file'])

            json_file = vehicle_base_name+".JSON"
            path = "./data/"+vehicle_file['folder']
            last_folder = os.path.basename(os.path.normpath(path))
            new_folder_path = os.path.join(output_dir, last_folder)
            if not os.path.exists(new_folder_path):
                os.makedirs(new_folder_path)
            name_folder_path = os.path.join(new_folder_path, row['full_name'])
            if not os.path.exists(name_folder_path):
                os.makedirs(name_folder_path)

            shutil.copy(path+"/"+json_file, name_folder_path+"/alt.json")
            shutil.copy(path+"/"+vehicle_base_name+".dds", name_folder_path+"/alt.dds")
            shutil.copy(path+"/"+vehicle_base_name+"_region.dds", name_folder_path+"/alt_region.dds")
            row['aaComponent'] = last_folder
            row['aaSkin'] = 'alt.dds'
        
            row['aaVehFile'] = vehicle_file['veh_file']
            result = template.format(**row)
            if path in counter:
                counter[path] = +1
            else:
                counter[path] = 0
            with open(new_folder_path+"/"+str(counter[path])+".rcd", "w") as f:
                f.write(result)


def createDataXlsx():
    workbook = openpyxl.Workbook()

    # Rename the default worksheet to "League"
    league_sheet = workbook.active
    league_sheet.title = "League"

   # Define the column widths for the "League" worksheet
    column_widths = [15, 15, 20, 8, 8, 20, 15, 8, 15, 15, 15, 15, 10, 15, 15, 15, 15, 15, 15]
    column_widths_liv = [25, 25, 40]

    # Add column headings to the "League" worksheet
    headings = ["first_name", "last_name", "full_name", "number", "season", "vehicle", "team",
                "speed", "qualify_speed", "wet_speed", "aggression", "composure", "crash",
                "completed_laps", "min_racing_skill", "start_skill", "recovery", "reputation",
                "courtesy"]
    league_sheet.append(headings)

    # Set the column widths and format the first row of the "League" worksheet
    for i, column_width in enumerate(column_widths):
        column_letter = openpyxl.utils.get_column_letter(i+1)
        league_sheet.column_dimensions[column_letter].width = column_width
        cell = league_sheet.cell(row=1, column=i+1)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

    # Add data to the "League" worksheet
    data = [
        ["Michael", "Schumacher", "Michael Schumacher", 3, 1, "Ferrari 488 GT3 1", "Ferrari",
        100, 100, 95, 100, 73, 100, 97, 48, 100, 80, 80, 80],
        ["Mika", "Häkkinen", "Mika Häkkinen", 1, 4, "AMG GT3 1", "Mercedes",
        100, 100, 94, 10, 69, 100, 95, 47, 100, 80, 80, 80],
        ["David", "Coulthard", "David Coulthard", 2, 8, "AMG GT3 2", "Mercedes",
        95, 95, 95, 90, 89, 100, 94, 45, 100, 80, 80, 80]
        # Add more data here...
    ]

    for row_data in data:
        league_sheet.append(row_data)

    # Add a new worksheet named "Liveries"
    liveries_sheet = workbook.create_sheet("Liveries")
    liveries_sheet.append(["livery_name", "folder", "veh_file"])

    # Add column headings to the "Liveries" worksheet
    for i, column_width in enumerate(column_widths_liv):
        column_letter = openpyxl.utils.get_column_letter(i+1)
        liveries_sheet.column_dimensions[column_letter].width = column_width
        cell = liveries_sheet.cell(row=1, column=i+1)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    # Add data to the "Liveries" worksheet
    liveries_sheet.append(["Ferrari 488 GT3 1", "Ferrari_488_GT3_2020", "F488GT3_0141079A6C.VEH"])
    liveries_sheet.append(["Ferrari 488 GT3 1", "Ferrari_488_GT3_2020", "F488GT3_51A3F2711F.VEH"])
    
    # Save the workbook
    workbook.save("data.xlsx")


#if data.xlsx exists create the roster
if os.path.isfile('data.xlsx'):
    createRoster() 
else:
    print("creating dummy data.xlsx")
    createDataXlsx()



