#import load_workbook Workbook os PatternFill
from openpyxl import load_workbook

import os
from openpyxl import Workbook
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.worksheet.datavalidation import DataValidation
#import PatternFill 
from openpyxl.styles import Font, PatternFill
import re

data_folder = './data'
mod_dir= 'C:/Program Files (x86)/Steam/steamapps/common/rFactor 2/Bin64/'
vehicles_dir = 'C:/Program Files (x86)/Steam/steamapps/common/rFactor 2/Installed/Vehicles'
mod_mgr = 'modmgr.exe'
name_filter = 'XXXX'
 

# Read XLSX file into Array
def readXLSX():
    wb = load_workbook(filename = 'data.xlsx')
    ws = wb.active
    data = []
    for row in ws.rows:
        data.append([cell.value for cell in row])
    return data


def getVehicles():
    vehicles = []
    files = os.listdir(vehicles_dir)
    for file_name in files:
        vehicles.append(file_name)
    return vehicles


def createConfigFile():
    vehicles = getVehicles()
    """create an XLSX file witht two columns: The first Column is the Vehicle Name and the Second is a Checbox if it should be selected"""
    wb = Workbook()
    ws = wb.active
    # set the headers for the columns
    ws.cell(row=1, column=1, value="Text").fill = PatternFill("solid", fgColor="0000FF")
    ws.cell(row=1, column=1).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=1, column=2, value="Checkbox").fill = PatternFill("solid", fgColor="0000FF")
    ws.cell(row=1, column=2).font = Font(bold=True, color="FFFFFF")
    dv = DataValidation(type="list", formula1='"True,False"')
    ws.add_data_validation(dv)
    # create an array of tupples with the first value from vehicles and the second value is FALSE. When the Name contains GT3 it is TRUE
    data = [(vehicle, "True" if name_filter in vehicle else "False") for vehicle in vehicles]
    print(data)
    for row_index, row in enumerate(data, start=2):
        text, value = row
        ws.cell(row=row_index, column=1, value=text)
        cell = ws.cell(row=row_index, column=2)
        cell.value = value
        dv.add(cell)

    # set the column width for the first column
    ws.column_dimensions['A'] = ColumnDimension(worksheet=ws, width=100)

    # set the column width for the second column
    ws.column_dimensions['B'] = ColumnDimension(worksheet=ws, width=100)
    # set the cell background color based on the value in the second column
    idx = 0
    for row in ws.iter_rows(min_row=2, max_row=9999, min_col=2, max_col=2):
        idx += 1
        for cell in row:
            if idx % 2 == 0:
                cell.offset(column=-1).fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
                cell.fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
            else:
                cell.offset(column=-1).fill = PatternFill(start_color="EECCEE", end_color="EECCEE", fill_type="solid")
                cell.fill = PatternFill(start_color="EECCEE", end_color="EECCEE", fill_type="solid")
    wb.save("config.xlsx")


def readConfigFile():
    wb = load_workbook('config.xlsx')
    # Select the active worksheet
    ws = wb.active
    # Initialize an empty list to store the values
    values = []
    # Loop through the rows in the worksheet
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Check if the second column is True
        if row[1] == "True":
            # If it is, add the value from the first column to the list
            values.append(vehicles_dir+"/"+row[0])

    latest_version_folder = []
    # Loop through each folder path in the array
    for folder_path in values:
        # Get a list of all subfolders in the current folder
        subfolders = next(os.walk(folder_path))[1]
        numeric_subfolders = [float(x) for x in subfolders]
        # Find the highest number in the list of numeric subfolders
        highest_number = max(numeric_subfolders, default=0)
        # Add the highest number to the list of highest numbers
        latest_version_folder.append(folder_path+"/"+str(highest_number))
    # Print the list of values
    for path in latest_version_folder:
        second_last_str = re.findall(r'/([^/]+)/[^/]+$', path)[0]
        subfolder_path = os.path.join(data_folder, second_last_str)
        if not os.path.exists(subfolder_path):
            os.makedirs(subfolder_path)
            out_subfolder_fullpath = os.getcwd().replace('\\', '/') + "/"+subfolder_path
            current = os.getcwd()
            os.chdir(mod_dir)
            os.system(mod_mgr+' *.veh *.dds *.json *288.png -x"'+path+"/"+"car-upgrade.mas"+'" -o"'+out_subfolder_fullpath+'"')
            os.chdir(current)



if not os.path.exists("config.xlsx"):
    print("config.xlsx does not exist")
    createConfigFile()
else:
    print("config.xlsx exists")
    readConfigFile()
